"""Esse módulo possui as funções exemplo de preenchimento de planilhas."""

from openpyxl import load_workbook

wb_example = load_workbook('Example.xlsx')
wb_conditions = load_workbook('Conditions.xlsx')

ws1_example = wb_example['Plan1']
ws1_conditions = wb_conditions['Plan1']


def value_in_cell(search, status):
    """Está função tem o objetivo de passar por todas as linhas de uma coluna
    e preencher com o status informado na coluna 'C' caso o valor seja encontrado dentro de uma célula.

    Parâmetros (search, status):
    Search: Valor a ser procurado dentro das células.
    Status: Status que será preenchido na coluna 'C'.
    """
    for cell in ws1_example['A']:
        if search in str(cell.value) and ws1_example['C' + str(cell.row) == None]:
            ws1_example['C' + str(cell.row)] = status
        elif search in str(cell.value) and ws1_example['C' + str(cell.row) != None]:
            print({search} + 'Value status populated with: ' + ws1_example['C' + str(cell.row)])


def value_is_equal(search, status):
    """Está função tem o objetivo de passar por todas as linhas de uma coluna
    e preencher com o status informado na coluna 'C' caso o valor encontrado seja idêntico ao da célula.

    Parâmetros (search, status):
    Search: Valor a ser procurado na coluna, identico ao da célula.
    Status: Status que será preenchido na coluna 'C'.
    """
    for cell in ws1_example['B']:    
        if cell.value == search and ws1_example['C' + str(cell.row)] == None:
            ws1_example['C' + str(cell.row)] = status


def equal_value_options(value1, value2, status):
    """Está função tem o objetivo de passar por todas as linhas de uma coluna
    e preencher com o status informado na coluna 'C' caso o valor encontrado seja idêntico ao da célula.
    Temos a opção de buscar dois valores e preencher com o mesmo status.

    Parâmetros (value1, value2, status):
    value1: Valor a ser procurado na coluna, identico ao da célula.
    value2: Segundo valor a ser procurado na coluna, identico ao da célula.
    Status: Status que será preenchido na coluna 'C'.
    """
    for cell in ws1_example['B']:
        if cell.value == value1 or cell.value == value2:
            ws1_example['C' + str(cell.row)] = status


def value_in_worksheet(status, column):
    """Está função tem o objetivo de buscar os dados informados em outra planilha,
    passar por todas as linhas de uma coluna de outra planilha
    e preencher com o status informado na coluna 'C' caso o valor encontrado seja idêntico ao da célula.

    Parâmetros (status, column):
    Status: Status que será preenchido na coluna 'C'.
    Column: Coluna que iremos buscar os dados na planilha que precisa ser preenchida.
    """
    count = 1
    value_in_column = "A1"
    while ws1_conditions.max_row >= count:
        for cell in ws1_example[column]:
            if str(ws1_conditions[value_in_column].value) == str(cell.value):
                row = 'C' + str(cell.row)
                value_in_column = 'A' + str(count)
                ws1_example[row] = status
            else:
                value_in_column = 'A' + str(count)
        count = count + 1

